# -*- coding: utf-8 -*-
"""
Created on Sun Aug  5 13:54:53 2018

@author: natsn
"""

import csv
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook
import openpyxl
import os

# Idea is to have all data processed into the dictionary, and then have the excel writer either
# write to csv, xlsx, append to, etc.
# Writes the internal buffered data chunk unless the user specifies an exact data chunk
# The SaveVert flag will flip data to be stored horizontally per key
class FileWriter:
    
    def __init__(self, path):
        self.path = path
        self.data_chunk = None
    
    # Reset where to save file
    def reset_path(self, new_path):
        self.path = new_path
    
    # Reset the internal chunk
    def clear_data_chunk(self):
        self.data_chunk = None
    
    # Append to the internal chunk
    def append_to_data_chunk(self, new_chunk):
        if self.data_chunk is None:
            self.assert_chunk_length(new_chunk)
            self.data_chunk = new_chunk
        else:
            assert list(self.data_chunk.keys()) == list(new_chunk.keys())
            for key in self.data_chunk.keys():
                assert len(new_chunk[key]) == len(self.data_chunk[key])
                self.data_chunk[key] += new_chunk[key]
    
    # Internal functions - asserts all dictionary lists are of same size       
    def assert_chunk_length(self, chunk):
        keys = list(chunk.keys())
        for i in range(len(keys) - 1):
            assert len(chunk[keys[i]]) == len(chunk[keys[i+1]])
    # Either write the current data chunk or write a specific one to the CSV file
    def write_csv(self, data_chunk = None, saveVert = True):
        if data_chunk is None:
            data_chunk = self.data_chunk
        self.assert_chunk_length(data_chunk)
        
        d = pd.DataFrame(data_chunk)
        if not saveVert:
            d = d.T
        d.to_csv(self.path, mode = 'w', header = True, index = False)
    
    # Append the current data chunk or append a new data chunk to a csv at self.path
    def write_append_csv(self, data_chunk = None, saveVert = True):
        if data_chunk is None:
            data_chunk = self.data_chunk
        self.assert_chunk_length(data_chunk)
        
        d = pd.DataFrame(data_chunk)
        if not saveVert:
            d = d.T
        d.to_csv(self.path, mode = 'a', header = False, index = False)
    
    # Writes the dictionary to an excel file instead
    def write_xlsx(self, data_chunk = None, saveVert = True):
        if data_chunk is None:
            data_chunk = self.data_chunk
        self.assert_chunk_length(data_chunk)
        
        a = pd.DataFrame(data_chunk)
        if not saveVert:
            a = a.T
        writer = pd.ExcelWriter(self.path, engine='xlsxwriter')
        a.to_excel(writer, sheet_name='Sheet1', index = False)
        writer.save()

    # Deletes still... no append
    def write_append_xlsx(self, data_chunk, saveVert = True):
        d = pd.DataFrame(data_chunk)
        if not saveVert:
            d = d.T
            
        with pd.ExcelWriter(self.path, engine='openpyxl') as writer:
            writer.book = load_workbook(self.path)
            mr = writer.book["Sheet1"].max_row
            d.to_excel(writer, "Sheet1", columns= list(data_chunk.keys()), startrow = mr)
            writer.save()



# Design Custom Pandas Excel Reader
data = {"does":[[0,0],0],"this":[1,1],"work":[2,2]}
csv_path = os.path.dirname(os.path.abspath(__file__)) + "\\test1.csv"
xlsx_path = os.path.dirname(os.path.abspath(__file__)) + "\\test2.xlsx"


xl = FileWriter(csv_path)
xl.append_to_data_chunk(data)
xl.append_to_data_chunk(data)
xl.write_csv(data)
xl.write_append_csv(data)
xl.reset_path(xlsx_path)
xl.write_xlsx(data)
xl.write_append_xlsx(data)

